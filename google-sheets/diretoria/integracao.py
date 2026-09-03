# -*- coding: utf-8 -*-
"""Compara os cabecalhos da planilha com os que o app escreve (google-sheets/Codigo.gs).
Se um nome divergir, o app preenche a aba e a apresentacao mostra zero."""
import re, json
from openpyxl import load_workbook

GS = open("/home/user/STRACTA-VIAGENS/google-sheets/Codigo.gs", encoding="utf-8").read()
bloco = GS[GS.index("var TABS = {"):GS.index("\n};", GS.index("var TABS = {"))]

abas = {}
for m in re.finditer(r'nome:\s*"([^"]+)",\s*dateCols:[^\n]*\n?(?:[^\n]*semId[^\n]*)?\s*colunas:\s*\[(.*?)\]',
                     bloco, re.S):
    nome, cols = m.group(1), m.group(2)
    abas[nome] = [c.strip().strip('"') for c in re.findall(r'"([^"]+)"', cols)]
sem_id = {m.group(1) for m in re.finditer(r'nome:\s*"([^"]+)"[^\n]*semId', bloco)}

import sys
wb = load_workbook(sys.argv[1] if len(sys.argv)>1 else "../GP2T-Planilha-Diretoria.xlsx")
falhas = 0
for nome, cols in abas.items():
    esperado = cols + ([] if nome in sem_id else ["_id"])
    if nome not in wb.sheetnames:
        print(f" FALHA aba '{nome}' não existe na planilha"); falhas += 1; continue
    ws = wb[nome]
    achado = [ws.cell(3, c).value for c in range(1, len(esperado) + 1)]
    if achado == esperado:
        print(f"  OK  {nome}: {len(esperado)} colunas conferem")
    else:
        falhas += 1
        print(f" FALHA {nome}")
        for i, (a, e) in enumerate(zip(achado, esperado), start=1):
            if a != e: print(f"        coluna {i}: planilha={a!r} · app={e!r}")
        if len(achado) != len(esperado):
            print(f"        quantidade: planilha={len(achado)} · app={len(esperado)}")

# as abas de apresentacao so podem ler as abas de resumo/cadastro, nunca dado cru
FONTES_OK = {"Resumo por Dia", "Resumo por Equipamento", "Resumo por Operador", "Resumo por Mês",
             "Equipamentos", "Operadores", "Manutenções", "Resumo Geral",
             "Por Equipamento", "Por Operador"}
for nome in ("Resumo Geral", "Por Equipamento", "Por Operador", "Evolução Diária",
             "Evolução Mensal", "Frota e Manutenção"):
    for row in wb[nome].iter_rows():
        for cel in row:
            if isinstance(cel.value, str) and cel.value.startswith("="):
                for ref in re.findall(r"'([^']+)'!", cel.value):
                    if ref not in FONTES_OK:
                        print(f" FALHA {nome}!{cel.coordinate} lê a aba '{ref}'"); falhas += 1
print("=== FALHAS DE INTEGRAÇÃO:", falhas, "===")
