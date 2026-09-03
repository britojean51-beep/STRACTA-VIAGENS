# -*- coding: utf-8 -*-
"""1) a planilha zerada abre sem erro de formula;
   2) ao receber dados (como o app escreve), os numeros aparecem."""
import datetime, shutil, formulas
from openpyxl import load_workbook

ARQ = "../GP2T-Planilha-Diretoria.xlsx"

def calcular(arq):
    sol = formulas.ExcelModel().loads(arq).finish().calculate()
    def val(chave):
        alvo = "]" + chave.upper().strip("'")
        for k, v in sol.items():
            if k.upper().endswith(alvo):
                try: return v.value[0, 0]
                except Exception: return v
    erros = []
    for k, v in sol.items():
        try: x = v.value[0, 0]
        except Exception: continue
        if isinstance(x, str) and x.startswith("#"): erros.append((k, x))
    return val, erros

log = []
def ok(nome, cond, extra=""):
    log.append(("  OK  " if cond else " FALHA ") + nome + (f" → {extra}" if extra != "" else ""))

# ---------- 1) vazia ----------
val, erros = calcular(ARQ)
ok("planilha vazia não tem NENHUMA célula de erro", len(erros) == 0,
   "; ".join(f"{k}={x}" for k, x in erros[:5]) or "0 erros")
for cel, nome in (("B7", "consumo"), ("B9", "produção"), ("B10", "viagens")):
    v = val(f"'Resumo Geral'!{cel}")
    ok(f"KPI {nome} zerado", (v in (0, 0.0, None)) or v == 0, v)
for cel, nome in (("B11", "L/h"), ("B12", "L/Ton"), ("B13", "km/L")):
    v = val(f"'Resumo Geral'!{cel}")
    ok(f"{nome} mostra travessão em vez de erro", str(v).strip() in ("—", "None", ""), v)
for cel, nome in (("B17", "melhor L/Ton"), ("B18", "pior L/Ton")):
    v = val(f"'Resumo Geral'!{cel}")
    ok(f"destaque {nome} mostra travessão", str(v).strip() in ("—", "None", ""), v)
ok("tabela de equipamentos vazia (sem #N/D)", str(val("'Por Equipamento'!A4")).strip() in ("", "None"),
   repr(val("'Por Equipamento'!A4")))

# ---------- 2) chegam os dados, como o app escreve ----------
shutil.copy(ARQ, "com-dados.xlsx")
wb = load_workbook("com-dados.xlsx")
D = lambda d: datetime.date(2026, 9, d)
# cadastros
for i, (eq, tipo) in enumerate([("CB-17", "Rodante (km/L)"), ("PC-01", "Horímetro (L/h)")]):
    r = 4 + i
    for c, v in enumerate([eq, tipo, "", "Operando", 1200 + i, 9000 + i, eq], start=1):
        wb["Equipamentos"].cell(r, c, v)
for i, op in enumerate(["Saulo", "José"]):
    for c, v in enumerate([op, "Operador", "Ativo", op], start=1):
        wb["Operadores"].cell(4 + i, c, v)
# resumo por dia: 2 dias
for i, (dia, diesel, horas, ton, viagens) in enumerate([(D(1), 300, 16, 400, 12), (D(2), 500, 18, 600, 20)]):
    for c, v in enumerate([dia, 2, 2, diesel, horas, diesel/horas, ton, diesel/ton,
                           diesel, 0, 0, 600, 600/diesel, viagens, 0, "CB-17, PC-01", "Saulo, José"], start=1):
        wb["Resumo por Dia"].cell(4 + i, c, v)
# resumo por equipamento
linhas_eq = [(D(1), "CB-17", 200, 8, 400, 600, 12), (D(1), "PC-01", 100, 8, 0, 0, 0),
             (D(2), "CB-17", 350, 9, 600, 600, 20), (D(2), "PC-01", 150, 9, 0, 0, 0)]
for i, (dia, eq, diesel, horas, ton, km, vi) in enumerate(linhas_eq):
    for c, v in enumerate([dia, eq, diesel, horas, diesel/horas, ton,
                           (diesel/ton if ton else 0), km, (km/diesel if diesel else 0), vi], start=1):
        wb["Resumo por Equipamento"].cell(4 + i, c, v)
# resumo por operador
for i, (dia, op, diesel, horas, ton, vi) in enumerate(
        [(D(1), "Saulo", 200, 8, 400, 12), (D(1), "José", 100, 8, 0, 0),
         (D(2), "Saulo", 350, 9, 600, 20), (D(2), "José", 150, 9, 0, 0)]):
    for c, v in enumerate([dia, op, "CB-17", diesel, horas, diesel/horas, ton,
                           (diesel/ton if ton else 0), vi], start=1):
        wb["Resumo por Operador"].cell(4 + i, c, v)
wb.save("com-dados.xlsx")

val, erros = calcular("com-dados.xlsx")
ok("com dados, ainda sem célula de erro", len(erros) == 0,
   "; ".join(f"{k}={x}" for k, x in erros[:5]) or "0 erros")
ok("consumo total = 800 L", float(val("'Resumo Geral'!B7")) == 800, val("'Resumo Geral'!B7"))
ok("produção total = 1.000 t", float(val("'Resumo Geral'!B9")) == 1000, val("'Resumo Geral'!B9"))
ok("L/Ton = 800/1000 = 0,80", abs(float(val("'Resumo Geral'!B12")) - 0.8) < 1e-9, val("'Resumo Geral'!B12"))
ok("equipamentos que operaram = 2", float(val("'Resumo Geral'!E7")) == 2, val("'Resumo Geral'!E7"))
ok("CB-17 apareceu sozinho na tabela", val("'Por Equipamento'!A4") == "CB-17", val("'Por Equipamento'!A4"))
ok("CB-17 · consumo 550 L", float(val("'Por Equipamento'!D4")) == 550, val("'Por Equipamento'!D4"))
ok("CB-17 · L/Ton = 550/1000 = 0,55", abs(float(val("'Por Equipamento'!H4")) - 0.55) < 1e-9, val("'Por Equipamento'!H4"))
ok("PC-01 (horímetro) · L/Ton mostra travessão",
   str(val("'Por Equipamento'!H5")).strip() == "—", val("'Por Equipamento'!H5"))
ok("Saulo apareceu em Por Operador", val("'Por Operador'!A4") == "Saulo", val("'Por Operador'!A4"))
ok("Saulo · produção 1.000 t", float(val("'Por Operador'!F4")) == 1000, val("'Por Operador'!F4"))
ok("Evolução Diária pegou o 1º dia", float(val("'Evolução Diária'!B4")) == 300, val("'Evolução Diária'!B4"))

print("=== ZERADA ==="); [print(l) for l in log]
print("=== FALHAS:", sum(1 for l in log if l.startswith(" FALHA")), "===")
