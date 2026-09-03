# -*- coding: utf-8 -*-
"""Planilha da Diretoria — apresentação, alimentada pelas abas de resumo do app."""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import sys
import dados

ZERADA = "--zerada" in sys.argv

FONTE = "Arial"
ESCURO = "0F172A"; AZUL = "1E3A8A"; CINZA = "E2E8F0"; CLARO = "F1F5F9"
VERDE = "047857"; VERMELHO = "B91C1C"; AMARELO = "FFF3B0"

def f(sz=10, b=False, cor="1F2937"): return Font(name=FONTE, size=sz, bold=b, color=cor)
def fill(c): return PatternFill("solid", fgColor=c)
BORDA = Border(*[Side(style="thin", color="CBD5E1")] * 4)

MAX = 3000           # até onde as fórmulas olham (sobra para anos de lançamento)
VAGAS = 30           # linhas reservadas nas tabelas: a frota pode crescer sem eu mexer
L1 = 4; L2 = 3 + VAGAS           # primeira e última linha de dados das tabelas
LTOT = L2 + 2                    # linha do total (fora dos intervalos de MIN/MAX/CONT)
CHART = 15                       # até onde os gráficos olham
RD = "'Resumo por Dia'"; RE = "'Resumo por Equipamento'"
RO = "'Resumo por Operador'"; RM = "'Resumo por Mês'"
ABA1 = "'Resumo Geral'"          # aba principal (o nome tem espaço: precisa de aspas nas fórmulas)
INI, FIM = f"{ABA1}!$B$4", f"{ABA1}!$D$4"

def periodo(aba, col_data):
    return f",{aba}!${col_data}$4:${col_data}${MAX},\">=\"&{INI},{aba}!${col_data}$4:${col_data}${MAX},\"<=\"&{FIM}"

def titulo(ws, texto, sub, largura):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=largura)
    ws.cell(1, 1, texto).font = f(16, True, "FFFFFF")
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    for c in range(1, largura + 1): ws.cell(1, c).fill = fill(ESCURO)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=largura)
    ws.cell(2, 1, sub).font = f(9, False, "64748B")

def cabecalho(ws, linha, cols, larguras):
    for i, (t, w) in enumerate(zip(cols, larguras), start=1):
        c = ws.cell(linha, i, t)
        c.font = f(10, True, "FFFFFF"); c.fill = fill(AZUL); c.border = BORDA
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[linha].height = 30

def zebra(ws, r1, r2, ncols):
    for r in range(r1, r2 + 1):
        for c in range(1, ncols + 1):
            cel = ws.cell(r, c); cel.border = BORDA
            if r % 2 == 0: cel.fill = fill(CLARO)

wb = Workbook()

# ==================== 1. RESUMO GERAL ====================
d = wb.active; d.title = "Resumo Geral"
titulo(d, "GP2T · GESTÃO DE FROTA", "Relatório da Diretoria — os números vêm do aplicativo, sem digitação.", 8)
if ZERADA:
    d["A3"] = "Planilha pronta para uso. Os números aparecem sozinhos assim que o app sincronizar (⚙️ Configurações → Sincronizar tudo)."
    d["A3"].font = f(9, True, VERDE)
else:
    d["A3"] = "⚠️ Este arquivo está com DADOS DE EXEMPLO (linhas com _id começando em EX-). Apague-as quando os dados reais entrarem."
    d["A3"].font = f(9, True, VERMELHO)

d["A4"] = "Período de:"; d["A4"].font = f(10, True)
# zerada abre no mês corrente; a de exemplo abre no período que tem dados
d["B4"] = datetime.date(2026, 9, 1) if ZERADA else datetime.date(2026, 8, 1)
d["D4"] = datetime.date(2026, 9, 30)
d["C4"] = "até:"; d["C4"].font = f(10, True)
d["E4"] = "Meta L/h:"; d["E4"].font = f(10, True); d["F4"] = 20
d["G4"] = "Meta km/L:"; d["G4"].font = f(10, True); d["H4"] = 1.0
for cel in ("B4", "D4", "F4", "H4"):
    d[cel].fill = fill(AMARELO); d[cel].font = f(10, True, "0000FF"); d[cel].border = BORDA
d["B4"].number_format = d["D4"].number_format = "DD/MM/YYYY"
d["H4"].number_format = "0.00"
d["A5"] = "As células amarelas são as únicas que você edita nesta aba."
d["A5"].font = f(8, False, "64748B")

KPIS = [
    ("Consumo total", f"=SUMIFS({RD}!$D$4:$D${MAX}{periodo(RD,'A')})", '#,##0" L"'),
    ("Horas trabalhadas", f"=SUMIFS({RD}!$E$4:$E${MAX}{periodo(RD,'A')})", '#,##0.0" h"'),
    ("Produção movida", f"=SUMIFS({RD}!$G$4:$G${MAX}{periodo(RD,'A')})", '#,##0" t"'),
    ("Viagens realizadas", f"=SUMIFS({RD}!$N$4:$N${MAX}{periodo(RD,'A')})", "#,##0"),
    ("Consumo por hora", '=IFERROR(B7/B8,"—")', '0.00" L/h"'),
    ("Consumo por tonelada", '=IFERROR(B7/B9,"—")', '0.00" L/t"'),
    ("Média da frota", f'=IFERROR(SUMIFS({RD}!$L$4:$L${MAX}{periodo(RD,"A")})/B7,"—")', '0.00" km/L"'),
    ("Dias com operação", f"=COUNTIFS({RD}!$A$4:$A${MAX},\">=\"&{INI},{RD}!$A$4:$A${MAX},\"<=\"&{FIM})", '#,##0" dias"'),
]
d["A6"] = "RESULTADO DO PERÍODO"; d["A6"].font = f(12, True, AZUL)
for i, (rot, fml, fmt) in enumerate(KPIS):
    r = 7 + i
    d.cell(r, 1, rot).font = f(11, True)
    c = d.cell(r, 2, fml); c.font = f(14, True, ESCURO); c.number_format = fmt
    c.alignment = Alignment(horizontal="left")
    d.cell(r, 1).border = BORDA; c.border = BORDA
d.column_dimensions["A"].width = 26; d.column_dimensions["B"].width = 18
for L, w in zip("CDEFGH", (8, 14, 12, 10, 12, 10)): d.column_dimensions[L].width = w

d["D6"] = "A FROTA"; d["D6"].font = f(12, True, AZUL)
FROTA = [
    ("Equipamentos que operaram", '=COUNTIF(\'Por Equipamento\'!$D$4:$D$33,">0")', "#,##0"),
    ("Operadores em atividade", '=COUNTIF(\'Por Operador\'!$C$4:$C$33,">0")', "#,##0"),
    ("Equipamentos cadastrados", f"=COUNTA(Equipamentos!$A$4:$A$200)", "#,##0"),
    ("Disponibilidade da frota", '=IFERROR(COUNTIF(Equipamentos!$D$4:$D$200,"Operando")/COUNTA(Equipamentos!$A$4:$A$200),"—")', "0%"),
    ("Manutenções no período", f"=SUMIFS({RD}!$O$4:$O${MAX}{periodo(RD,'A')})", "#,##0"),
    ("Acima da meta de L/h", '=COUNTIF(\'Por Equipamento\'!$F$4:$F$33,">"&$F$4)', "#,##0"),
]
for i, (rot, fml, fmt) in enumerate(FROTA):
    r = 7 + i
    d.cell(r, 4, rot).font = f(11, True); d.cell(r, 4).border = BORDA
    c = d.cell(r, 5, fml); c.font = f(14, True, ESCURO); c.number_format = fmt; c.border = BORDA

d["A16"] = "DESTAQUES"; d["A16"].font = f(12, True, AZUL)
DEST = [
    ("Menor consumo por tonelada",
     '=IFERROR(INDEX(\'Por Equipamento\'!$A$4:$A$33,MATCH(MIN(\'Por Equipamento\'!$H$4:$H$33),\'Por Equipamento\'!$H$4:$H$33,0)),"—")',
     '=IFERROR(MIN(\'Por Equipamento\'!$H$4:$H$33),"—")', '0.00" L/t"'),
    ("Maior consumo por tonelada",
     '=IFERROR(INDEX(\'Por Equipamento\'!$A$4:$A$33,MATCH(MAX(\'Por Equipamento\'!$H$4:$H$33),\'Por Equipamento\'!$H$4:$H$33,0)),"—")',
     '=IFERROR(MAX(\'Por Equipamento\'!$H$4:$H$33),"—")', '0.00" L/t"'),
    ("Operador com mais produção",
     '=IFERROR(INDEX(\'Por Operador\'!$A$4:$A$33,MATCH(MAX(\'Por Operador\'!$F$4:$F$33),\'Por Operador\'!$F$4:$F$33,0)),"—")',
     '=IFERROR(MAX(\'Por Operador\'!$F$4:$F$33),"—")', '#,##0" t"'),
    ("Equipamento que mais consumiu",
     '=IFERROR(INDEX(\'Por Equipamento\'!$A$4:$A$33,MATCH(MAX(\'Por Equipamento\'!$D$4:$D$33),\'Por Equipamento\'!$D$4:$D$33,0)),"—")',
     '=IFERROR(MAX(\'Por Equipamento\'!$D$4:$D$33),"—")', '#,##0" L"'),
]
for i, (rot, quem, valor, fmt) in enumerate(DEST):
    r = 17 + i
    d.cell(r, 1, rot).font = f(10, True); d.cell(r, 1).border = BORDA
    c = d.cell(r, 2, quem); c.font = f(12, True, VERDE if i in (0, 2) else VERMELHO); c.border = BORDA
    v = d.cell(r, 3, valor); v.font = f(11, True); v.number_format = fmt; v.border = BORDA

d["A23"] = "COMO ESTA PLANILHA SE MANTÉM ATUALIZADA"; d["A23"].font = f(12, True, AZUL)
PASSOS = [
    "1. Suba este arquivo no Google Drive e abra como Planilhas Google (o app não escreve em .xlsx).",
    "2. No Apps Script do app, coloque o link desta planilha em PLANILHA_URL e faça Implantar → Nova versão.",
    "3. No aplicativo: ⚙️ Configurações → 🔄 Sincronizar tudo.",
    "4. Apague as linhas de exemplo (coluna _id começando com EX-) nas abas de dados.",
    "",
    "As abas cinzas (Resumo por Dia, Abastecimentos, Viagens…) são preenchidas pelo aplicativo a cada lançamento.",
    "As abas azuis são só apresentação: leem as abas cinzas e não recalculam nada por conta própria —",
    "por isso os números aqui são sempre iguais aos que aparecem no celular.",
]
for i, t in enumerate(PASSOS):
    d.cell(24 + i, 1, t).font = f(9, i >= 5, "1F2937" if i < 5 else AZUL)
d.freeze_panes = "A6"
d.sheet_view.showGridLines = False

# ==================== 2. POR EQUIPAMENTO ====================
pe = wb.create_sheet("Por Equipamento")
titulo(pe, "DESEMPENHO POR EQUIPAMENTO", "Período definido na aba Resumo Geral. Fonte: aba Resumo por Equipamento (preenchida pelo app).", 11)
COLS = ["Equipamento", "Tipo", "Dias operados", "Consumo (L)", "Horas", "L/h",
        "Produção (t)", "L/Ton", "KM rodado", "Média km/L", "Viagens"]
cabecalho(pe, 3, COLS, [14, 18, 12, 13, 10, 10, 13, 10, 12, 12, 10])
for i in range(VAGAS):
    r = L1 + i; o = 4 + i          # linha correspondente na aba Equipamentos
    vazio = f'Equipamentos!$A{o}=""'
    p = periodo(RE, "A")
    crit = f",{RE}!$B$4:$B${MAX},$A{r}"
    pe.cell(r, 1, f'=IF({vazio},"",Equipamentos!$A{o})').font = f(10, True)
    pe.cell(r, 2, f'=IF({vazio},"",Equipamentos!$B{o})')
    pe.cell(r, 3, f'=IF({vazio},"",COUNTIFS({RE}!$B$4:$B${MAX},$A{r}{p},{RE}!$C$4:$C${MAX},">0"))')
    pe.cell(r, 4, f'=IF({vazio},"",SUMIFS({RE}!$C$4:$C${MAX}{crit}{p}))')
    pe.cell(r, 5, f'=IF({vazio},"",SUMIFS({RE}!$D$4:$D${MAX}{crit}{p}))')
    pe.cell(r, 6, f'=IF({vazio},"",IFERROR(IF(E{r}=0,"—",D{r}/E{r}),"—"))')
    pe.cell(r, 7, f'=IF({vazio},"",SUMIFS({RE}!$F$4:$F${MAX}{crit}{p}))')
    pe.cell(r, 8, f'=IF({vazio},"",IFERROR(IF(G{r}=0,"—",D{r}/G{r}),"—"))')
    pe.cell(r, 9, f'=IF({vazio},"",SUMIFS({RE}!$H$4:$H${MAX}{crit}{p}))')
    pe.cell(r, 10, f'=IF({vazio},"",IFERROR(IF(OR(D{r}=0,I{r}=0),"—",I{r}/D{r}),"—"))')
    pe.cell(r, 11, f'=IF({vazio},"",SUMIFS({RE}!$J$4:$J${MAX}{crit}{p}))')
    for c, fmt in ((3, "#,##0"), (4, "#,##0"), (5, "#,##0.0"), (6, "0.00"), (7, "#,##0"),
                   (8, "0.00"), (9, "#,##0"), (10, "0.00"), (11, "#,##0")):
        pe.cell(r, c).number_format = fmt
ult = L2
zebra(pe, 4, ult, 11)
pe.cell(LTOT, 1, "TOTAL DA FROTA").font = f(10, True)
for c in (3, 4, 5, 7, 9, 11):
    L = get_column_letter(c)
    pe.cell(LTOT, c, f"=SUM({L}{L1}:{L}{L2})").font = f(10, True)
pe.cell(LTOT, 6, f'=IFERROR(D{LTOT}/E{LTOT},"—")').font = f(10, True)
pe.cell(LTOT, 8, f'=IFERROR(D{LTOT}/G{LTOT},"—")').font = f(10, True)
pe.cell(LTOT, 10, f'=IFERROR(I{LTOT}/D{LTOT},"—")').font = f(10, True)
for c, fmt in ((1, None), (2, None), (3, "#,##0"), (4, "#,##0"), (5, "#,##0.0"), (6, "0.00"),
               (7, "#,##0"), (8, "0.00"), (9, "#,##0"), (10, "0.00"), (11, "#,##0")):
    cel = pe.cell(LTOT, c); cel.fill = fill(CINZA); cel.border = BORDA; cel.font = f(10, True)
    if fmt: cel.number_format = fmt
pe.freeze_panes = "A4"; pe.sheet_view.showGridLines = False

g1 = BarChart(); g1.title = "Consumo de diesel por equipamento (L)"; g1.style = 10
g1.add_data(Reference(pe, min_col=4, min_row=3, max_row=CHART), titles_from_data=True)
g1.set_categories(Reference(pe, min_col=1, min_row=4, max_row=CHART))
g1.height, g1.width, g1.legend = 8, 18, None
pe.add_chart(g1, f"A{LTOT + 3}")
g2 = BarChart(); g2.title = "Consumo por tonelada movida (L/t) — quanto menor, melhor"; g2.style = 12
g2.add_data(Reference(pe, min_col=8, min_row=3, max_row=CHART), titles_from_data=True)
g2.set_categories(Reference(pe, min_col=1, min_row=4, max_row=CHART))
g2.height, g2.width, g2.legend = 8, 18, None
pe.add_chart(g2, f"A{LTOT + 21}")

# ==================== 3. POR OPERADOR ====================
po = wb.create_sheet("Por Operador")
titulo(po, "DESEMPENHO POR OPERADOR", "Período definido na aba Resumo Geral. Fonte: aba Resumo por Operador (preenchida pelo app).", 8)
COLS_O = ["Operador", "Dias trabalhados", "Consumo (L)", "Horas", "L/h", "Produção (t)", "L/Ton", "Viagens"]
cabecalho(po, 3, COLS_O, [16, 14, 13, 10, 10, 13, 10, 10])
p = periodo(RO, "A")
for i in range(VAGAS):
    r = L1 + i; o = 4 + i
    vazio = f'Operadores!$A{o}=""'
    crit = f",{RO}!$B$4:$B${MAX},$A{r}"
    po.cell(r, 1, f'=IF({vazio},"",Operadores!$A{o})').font = f(10, True)
    po.cell(r, 2, f'=IF({vazio},"",COUNTIFS({RO}!$B$4:$B${MAX},$A{r}{p}))')
    po.cell(r, 3, f'=IF({vazio},"",SUMIFS({RO}!$D$4:$D${MAX}{crit}{p}))')
    po.cell(r, 4, f'=IF({vazio},"",SUMIFS({RO}!$E$4:$E${MAX}{crit}{p}))')
    po.cell(r, 5, f'=IF({vazio},"",IFERROR(IF(D{r}=0,"—",C{r}/D{r}),"—"))')
    po.cell(r, 6, f'=IF({vazio},"",SUMIFS({RO}!$G$4:$G${MAX}{crit}{p}))')
    po.cell(r, 7, f'=IF({vazio},"",IFERROR(IF(F{r}=0,"—",C{r}/F{r}),"—"))')
    po.cell(r, 8, f'=IF({vazio},"",SUMIFS({RO}!$I$4:$I${MAX}{crit}{p}))')
    for c, fmt in ((2, "#,##0"), (3, "#,##0"), (4, "#,##0.0"), (5, "0.00"),
                   (6, "#,##0"), (7, "0.00"), (8, "#,##0")):
        po.cell(r, c).number_format = fmt
ult_o = L2
zebra(po, 4, ult_o, 8)
po.freeze_panes = "A4"; po.sheet_view.showGridLines = False
g3 = BarChart(); g3.title = "Produção movida por operador (t)"; g3.style = 11
g3.add_data(Reference(po, min_col=6, min_row=3, max_row=CHART), titles_from_data=True)
g3.set_categories(Reference(po, min_col=1, min_row=4, max_row=CHART))
g3.height, g3.width, g3.legend = 8, 18, None
po.add_chart(g3, f"A{ult_o + 3}")

# ==================== 4. EVOLUÇÃO DIÁRIA ====================
ev = wb.create_sheet("Evolução Diária")
titulo(ev, "EVOLUÇÃO DIÁRIA", "Espelho da aba Resumo por Dia — cresce sozinha conforme o app lança.", 8)
cabecalho(ev, 3, ["Data", "Consumo (L)", "Horas", "L/h", "Produção (t)", "L/Ton", "Viagens", "Equip."],
          [12, 13, 10, 10, 13, 10, 10, 10])
LINHAS_DIA = 200
for i in range(LINHAS_DIA):
    r, o = 4 + i, 4 + i
    ev.cell(r, 1, f'=IF({RD}!$A{o}="","",{RD}!$A{o})').number_format = "DD/MM/YYYY"
    for c, src, fmt in ((2, "D", "#,##0"), (3, "E", "#,##0.0"), (4, "F", "0.00"),
                        (5, "G", "#,##0"), (6, "H", "0.00"), (7, "N", "#,##0"), (8, "B", "#,##0")):
        ev.cell(r, c, f'=IF({RD}!$A{o}="","",{RD}!${src}{o})').number_format = fmt
ev.freeze_panes = "A4"; ev.sheet_view.showGridLines = False
g4 = LineChart(); g4.title = "Consumo (L) x Produção (t) por dia"; g4.style = 12
g4.add_data(Reference(ev, min_col=2, min_row=3, max_row=3 + 40), titles_from_data=True)
g4.add_data(Reference(ev, min_col=5, min_row=3, max_row=3 + 40), titles_from_data=True)
g4.set_categories(Reference(ev, min_col=1, min_row=4, max_row=3 + 40))
g4.height, g4.width = 9, 22
ev.add_chart(g4, "J3")
g5 = LineChart(); g5.title = "Consumo por tonelada (L/t) por dia"; g5.style = 13
g5.add_data(Reference(ev, min_col=6, min_row=3, max_row=3 + 40), titles_from_data=True)
g5.set_categories(Reference(ev, min_col=1, min_row=4, max_row=3 + 40))
g5.height, g5.width, g5.legend = 9, 22, None
ev.add_chart(g5, "J22")

# ==================== 5. EVOLUÇÃO MENSAL ====================
em = wb.create_sheet("Evolução Mensal")
titulo(em, "FECHAMENTO MENSAL", "Espelho da aba Resumo por Mês — uma linha por mês, escrita pelo app.", 9)
cabecalho(em, 3, ["Mês", "Dias", "Consumo (L)", "Horas", "L/h", "Produção (t)", "L/Ton", "Viagens", "Manut."],
          [16, 8, 13, 10, 10, 13, 10, 10, 9])
for i in range(36):
    r, o = 4 + i, 4 + i
    em.cell(r, 1, f'=IF({RM}!$A{o}="","",{RM}!$A{o})')
    for c, src, fmt in ((2, "B", "#,##0"), (3, "E", "#,##0"), (4, "F", "#,##0.0"), (5, "G", "0.00"),
                        (6, "H", "#,##0"), (7, "I", "0.00"), (8, "O", "#,##0"), (9, "P", "#,##0")):
        em.cell(r, c, f'=IF({RM}!$A{o}="","",{RM}!${src}{o})').number_format = fmt
em.freeze_panes = "A4"; em.sheet_view.showGridLines = False
g6 = BarChart(); g6.title = "Consumo mensal (L)"; g6.style = 10
g6.add_data(Reference(em, min_col=3, min_row=3, max_row=15), titles_from_data=True)
g6.set_categories(Reference(em, min_col=1, min_row=4, max_row=15))
g6.height, g6.width, g6.legend = 8, 20, None
em.add_chart(g6, "K3")

# ==================== 6. FROTA E MANUTENÇÃO ====================
fm = wb.create_sheet("Frota e Manutenção")
titulo(fm, "FROTA E MANUTENÇÃO", "Situação atual de cada equipamento. Fonte: abas Equipamentos e Manutenções.", 7)
cabecalho(fm, 3, ["Equipamento", "Tipo", "Situação", "Horímetro atual", "KM atual",
                  "Manutenções no período", "Última manutenção"], [14, 18, 14, 15, 13, 16, 16])
for i in range(VAGAS):
    r = L1 + i; o = 4 + i
    vazio = f'Equipamentos!$A{o}=""'
    fm.cell(r, 1, f'=IF({vazio},"",Equipamentos!$A{o})').font = f(10, True)
    fm.cell(r, 2, f'=IF({vazio},"",Equipamentos!$B{o})')
    fm.cell(r, 3, f'=IF({vazio},"",Equipamentos!$D{o})')
    fm.cell(r, 4, f'=IF({vazio},"",Equipamentos!$E{o})').number_format = "#,##0.0"
    fm.cell(r, 5, f'=IF({vazio},"",Equipamentos!$F{o})').number_format = "#,##0"
    fm.cell(r, 6, f'=IF({vazio},"",COUNTIFS(Manutenções!$B$4:$B$500,$A{r},Manutenções!$A$4:$A$500,">="&{INI},'
                  f'Manutenções!$A$4:$A$500,"<="&{FIM}))').number_format = "#,##0"
    fm.cell(r, 7, f'=IF({vazio},"",IF(SUMPRODUCT(MAX((Manutenções!$B$4:$B$500=$A{r})*(Manutenções!$A$4:$A$500)))=0,"—",'
                  f'SUMPRODUCT(MAX((Manutenções!$B$4:$B$500=$A{r})*(Manutenções!$A$4:$A$500)))))').number_format = "DD/MM/YYYY"
ult_f = L2
zebra(fm, 4, ult_f, 7)
fm.cell(ult_f + 2, 1, "MANUTENÇÕES REGISTRADAS").font = f(12, True, AZUL)
cabecalho(fm, ult_f + 3, ["Data", "Equipamento", "Responsável", "Tipo", "Serviço realizado", "Peças/Trocas", "Observação"],
          [12, 14, 16, 14, 30, 26, 24])
for i in range(60):
    r, o = ult_f + 4 + i, 4 + i
    fm.cell(r, 1, f'=IF(Manutenções!$A{o}="","",Manutenções!$A{o})').number_format = "DD/MM/YYYY"
    for c, src in ((2, "B"), (3, "C"), (4, "F"), (5, "G"), (6, "H"), (7, "I")):
        fm.cell(r, c, f'=IF(Manutenções!$A{o}="","",Manutenções!${src}{o})')
fm.freeze_panes = "A4"; fm.sheet_view.showGridLines = False

# ==================== ABAS DO APP ====================
abast, viagens, manut, estado = dados.gerar()
por_dia, por_eq, por_op, por_mes = dados.resumos(abast, viagens, manut)
if ZERADA:                       # só o cabeçalho: o app escreve a partir da linha 4
    abast = viagens = manut = por_dia = por_eq = por_op = por_mes = []
    estado = {}
DT = lambda iso: datetime.date(*map(int, iso.split("-")))

def aba_app(nome, nota, cols, larguras, linhas):
    ws = wb.create_sheet(nome)
    ws.cell(1, 1, nome).font = f(13, True, ESCURO)
    ws.cell(2, 1, nota).font = f(9, False, "64748B")
    cabecalho(ws, 3, cols, larguras)
    for i, linha in enumerate(linhas):
        for j, v in enumerate(linha, start=1):
            c = ws.cell(4 + i, j, v)
            c.font = f(9); c.border = BORDA
            if isinstance(v, datetime.date): c.number_format = "DD/MM/YYYY"
            elif isinstance(v, float): c.number_format = "#,##0.00"
    ws.freeze_panes = "A4"
    return ws

NOTA = "Preenchido automaticamente pelo app GP2T — não edite à mão."
aba_app("Resumo por Dia", NOTA,
        ["Data", "Equipamentos", "Operadores", "Consumo total (L)", "Horas totais", "L/h", "Produção (t)",
         "L/Ton", "Diesel S-10 (L)", "Diesel S-500 (L)", "ARLA (L)", "KM", "Média km/L", "Viagens",
         "Manutenções", "Quais equipamentos", "Quais operadores"],
        [12, 12, 11, 14, 12, 9, 12, 9, 13, 14, 11, 10, 12, 10, 11, 40, 30],
        [[DT(x["data"]), x["equipamentos"], x["operadores"], x["diesel"], x["horas"], round(x["lh"], 2),
          x["ton"], round(x["lton"], 2), x["s10"], x["s500"], x["arla"], x["km"], round(x["media"], 2),
          x["viagens"], x["manutencoes"], x["quais_eq"], x["quais_op"]] for x in por_dia])
aba_app("Resumo por Equipamento", NOTA,
        ["Data", "Equipamento", "Consumo (L)", "Horas", "L/h", "Produção (t)", "L/Ton", "KM", "Média km/L", "Viagens"],
        [12, 14, 12, 10, 9, 12, 9, 10, 12, 10],
        [[DT(x["data"]), x["equipamento"], x["diesel"], x["horas"], round(x["lh"], 2), x["ton"],
          round(x["lton"], 2), x["km"], round(x["media"], 2), x["viagens"]] for x in por_eq])
aba_app("Resumo por Operador", NOTA,
        ["Data", "Operador", "Equipamentos", "Consumo (L)", "Horas", "L/h", "Produção (t)", "L/Ton", "Viagens"],
        [12, 14, 22, 12, 10, 9, 12, 9, 10],
        [[DT(x["data"]), x["operador"], x["equipamentos"], x["diesel"], x["horas"], round(x["lh"], 2),
          x["ton"], round(x["lton"], 2), x["viagens"]] for x in por_op])
aba_app("Resumo por Mês", NOTA,
        ["Mês", "Dias com lançamento", "Equipamentos", "Operadores", "Consumo total (L)", "Horas totais",
         "L/h", "Produção (t)", "L/Ton", "Diesel S-10 (L)", "Diesel S-500 (L)", "ARLA (L)", "KM",
         "Média km/L", "Viagens", "Manutenções", "Quais equipamentos", "Quais operadores", "_id"],
        [16, 14, 12, 11, 14, 12, 9, 12, 9, 13, 14, 11, 10, 12, 10, 11, 40, 30, 12],
        [[x["mes"], x["dias"], x["equipamentos"], x["operadores"], x["diesel"], x["horas"], round(x["lh"], 2),
          x["ton"], round(x["lton"], 2), x["s10"], x["s500"], x["arla"], x["km"], round(x["media"], 2),
          x["viagens"], x["manutencoes"], x["quais_eq"], x["quais_op"], x["chave"]] for x in por_mes])
aba_app("Abastecimentos", NOTA,
        ["Data", "Equipamento", "Operador", "Horímetro Inicial", "Horímetro Final", "Horas", "Litros",
         "Combustível", "ARLA (L)", "KM Rodado", "Média", "Unidade", "Toneladas", "L/Ton", "Situação", "_id"],
        [12, 13, 12, 14, 14, 8, 9, 12, 10, 11, 9, 9, 11, 9, 20, 12],
        [[DT(a["dia"]), a["equipamento"], a["operador"], a["hor_ini"], a["hor_fim"], a["horas"], a["litros"],
          a["combustivel"], a["arla"], a["km"], a["media"], a["unidade"], a["toneladas"] or "", a["lton"] or "",
          a["situacao"], a["id"]] for a in abast])
aba_app("Viagens", NOTA,
        ["Data", "Equipamento", "Operador", "Origem", "Destino", "Viagens", "Material",
         "Peso/viagem (t)", "Peso total (t)", "_id"],
        [12, 13, 12, 9, 9, 9, 12, 14, 14, 12],
        [[DT(v["dia"]), v["equipamento"], v["operador"], v["origem"], v["destino"], v["quantidade"],
          v["material"], v["peso_viagem"], v["peso_total"], v["id"]] for v in viagens])
aba_app("Manutenções", NOTA,
        ["Data", "Equipamento", "Operador/Responsável", "Horímetro", "KM", "Tipo", "Serviço Realizado",
         "Peças/Trocas", "Observação", "Próxima Manutenção", "_id"],
        [12, 13, 18, 11, 10, 13, 28, 26, 24, 16, 12],
        [[DT(m["dia"]), m["equipamento"], m["responsavel"], "", "", m["tipo"], m["servico"],
          m["pecas"], m["obs"], "", m["id"]] for m in manut])
aba_app("Equipamentos", NOTA,
        ["Código", "Tipo", "Modelo", "Status", "Horímetro Atual", "KM Atual", "_id"],
        [12, 18, 16, 14, 14, 12, 12],
        [] if ZERADA else
        [[eq, tipo, "", "Manutenção" if eq == "PC-02" else "Operando",
          round(estado.get(eq, {}).get("hor", 0), 1), estado.get(eq, {}).get("km", 0), eq]
         for eq, tipo in dados.EQUIPS])
aba_app("Operadores", NOTA, ["Nome", "Função", "Status", "_id"], [18, 14, 12, 14],
        [] if ZERADA else [[o, "Operador", "Ativo", o] for o in dados.OPERADORES])

# cor das guias: azul = apresentação, cinza = alimentada pelo app
for nome in ("Resumo Geral", "Por Equipamento", "Por Operador", "Evolução Diária", "Evolução Mensal", "Frota e Manutenção"):
    wb[nome].sheet_properties.tabColor = AZUL
for nome in ("Resumo por Dia", "Resumo por Equipamento", "Resumo por Operador", "Resumo por Mês",
             "Abastecimentos", "Viagens", "Manutenções", "Equipamentos", "Operadores"):
    wb[nome].sheet_properties.tabColor = "94A3B8"

SAIDA = ("/home/user/STRACTA-VIAGENS/google-sheets/GP2T-Planilha-Diretoria.xlsx" if ZERADA else
         "/home/user/STRACTA-VIAGENS/google-sheets/GP2T-Planilha-Diretoria-EXEMPLO.xlsx")
wb.save(SAIDA)
print("salvo:", SAIDA, "| abas:", len(wb.sheetnames))
